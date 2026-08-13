#!/usr/bin/env python3
"""Build GovHub_SaaS_Metrics_v2.xlsx - honest, real-data SaaS workbook."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

OUT = "/tmp/claude-0/-home-user/d0066b2f-6ca2-5e2b-baa6-092f3522a712/scratchpad/GovHub_SaaS_Metrics_v2.xlsx"

FONT = "Arial"
BLUE   = Font(name=FONT, size=10, color="0000FF")            # hardcoded input
BLACK  = Font(name=FONT, size=10, color="000000")            # formula
GREEN  = Font(name=FONT, size=10, color="008000")            # cross-sheet link
BOLD   = Font(name=FONT, size=10, bold=True)
H1     = Font(name=FONT, size=14, bold=True, color="FFFFFF")
H2     = Font(name=FONT, size=11, bold=True, color="1F3864")
NOTE   = Font(name=FONT, size=9, italic=True, color="595959")
SMALL  = Font(name=FONT, size=9, color="000000")
WARN   = Font(name=FONT, size=10, bold=True, color="C00000")

FILL_H1   = PatternFill("solid", fgColor="1F3864")
FILL_H2   = PatternFill("solid", fgColor="D9E2F3")
FILL_YEL  = PatternFill("solid", fgColor="FFFF00")
FILL_GREY = PatternFill("solid", fgColor="F2F2F2")
FILL_RED  = PatternFill("solid", fgColor="FCE4E4")
FILL_GRN  = PatternFill("solid", fgColor="E2EFDA")

CUR   = '$#,##0;($#,##0);-'
CUR2  = '$#,##0.00;($#,##0.00);-'
PCT   = '0.0%;(0.0%);-'
PCT2  = '0.00%;(0.00%);-'
NUM   = '#,##0;(#,##0);-'
NUM1  = '#,##0.0;(#,##0.0);-'
MULT  = '0.0x'

thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

# ---------------------------------------------------------------- actual data
# 12 months of history: Sep-25 .. Aug-26. Aug-26 is month-to-date (through 8/13).
MONTHS = ["Sep-25","Oct-25","Nov-25","Dec-25","Jan-26","Feb-26",
          "Mar-26","Apr-26","May-26","Jun-26","Jul-26","Aug-26"]
NM = len(MONTHS)
C0 = 3                      # first month column (C)
CL = [get_column_letter(C0+i) for i in range(NM)]
TOTC = get_column_letter(C0+NM)     # totals column (O)

# Supabase GovProp (project iqekrwearenblsmhvdjn), queried 2026-08-13
SIGNUPS      = [1, 11, 1, 1, 0, 0, 1, 0, 0, 0, 12, 3]     # user_profiles, deleted_at is null
SIGNUPS_EXT  = [1,  2, 1, 1, 0, 0, 1, 0, 0, 0,  3, 0]     # non-founder / non-test email
SIGNUPS_INT  = [0,  9, 0, 0, 0, 0, 0, 0, 0, 0,  9, 3]     # founder aliases + test accounts
ONBOARDED    = [1,  7, 0, 0, 0, 0, 0, 0, 0, 0, 11, 2]     # onboarding_completed = true
SUBMISSIONS  = [0,  0, 0, 0, 0, 0, 0, 0, 6, 3, 16, 5]     # v2_submissions
API_COST     = [0,0,0,0,0,0,0,0,0, 0.19, 26.75, 7.90]     # v2_token_usage.est_cost_usd
LIFECYCLE_EM = [0,0,0,0,0,0,0,0,0, 19, 65, 6]             # scheduled_emails status='sent'

# Semrush (us database), pulled 2026-08-13
SEO_KW       = [0,0,0,1,1,0,0,0,1,1,90,114]               # organic keywords in top 100
SEO_TRAFFIC  = [0,0,0,0,0,0,0,0,0,0,21,20]                # est. monthly organic sessions
GSC_CLICKS   = [0,0,0,0,0,0,0,0,0,0,15,None]              # GSC 28-day trailing (Jul 24 milestone)

# Cold email - Instantly wave 1 (docs/instantly-wave1.md), campaigns still in Draft
EMAILS_SENT  = [0]*12
CE_DOMAINS   = [0,0,0,0,0,0,0,0,0,0,0,16]
CE_MAILBOXES = [0,0,0,0,0,0,0,0,0,0,0,12]

# Vendor / PR outreach via Resend (govhub-marketing data/vendor-outreach.json)
VENDOR_SENT  = [0,0,0,0,0,0,0,0,0,0,10,38]

wb = openpyxl.Workbook()


def sheet(name):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    return ws


def title(ws, text, sub, width):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    c = ws.cell(1, 1, text); c.font = H1; c.fill = FILL_H1
    c.alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 24
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    c = ws.cell(2, 1, sub); c.font = NOTE
    c.alignment = Alignment(vertical="center", indent=1, wrap_text=False)
    ws.row_dimensions[2].height = 14


def band(ws, row, text, width):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    c = ws.cell(row, 1, text); c.font = H2; c.fill = FILL_H2
    c.alignment = Alignment(vertical="center", indent=1)
    return row + 1


def monthhdr(ws, row, label="Metric", note="Source / note"):
    ws.cell(row, 1, label).font = BOLD
    ws.cell(row, 2, note).font = BOLD
    for i, m in enumerate(MONTHS):
        c = ws.cell(row, C0+i, m); c.font = BOLD
        c.alignment = Alignment(horizontal="center")
        c.fill = FILL_GREY
    c = ws.cell(row, C0+NM, "TOTAL / NOW"); c.font = BOLD
    c.alignment = Alignment(horizontal="center"); c.fill = FILL_GREY
    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 40
    for i in range(NM+1):
        ws.column_dimensions[get_column_letter(C0+i)].width = 10.5
    return row + 1


def datarow(ws, row, label, note, vals, fmt=NUM, total="sum", font=BLUE, indent=0):
    c = ws.cell(row, 1, label); c.font = BOLD if indent == 0 else BLACK
    c.alignment = Alignment(indent=indent)
    ws.cell(row, 2, note).font = NOTE
    for i, v in enumerate(vals):
        cc = ws.cell(row, C0+i)
        if v is not None:
            cc.value = v
        cc.font = font; cc.number_format = fmt
    tc = ws.cell(row, C0+NM)
    if total == "sum":
        tc.value = f"=SUM({CL[0]}{row}:{CL[-1]}{row})"
    elif total == "last":
        tc.value = f"={CL[-1]}{row}"
    elif total is None:
        pass
    else:
        tc.value = total
    tc.font = BOLD; tc.number_format = fmt; tc.fill = FILL_GREY
    return row + 1


def formularow(ws, row, label, note, f_template, fmt=PCT, total=None, indent=1):
    """f_template uses {c} for the month column letter."""
    c = ws.cell(row, 1, label); c.font = BLACK
    c.alignment = Alignment(indent=indent)
    ws.cell(row, 2, note).font = NOTE
    for i in range(NM):
        cc = ws.cell(row, C0+i, f_template.format(c=CL[i]))
        cc.font = BLACK; cc.number_format = fmt
    tc = ws.cell(row, C0+NM)
    if total:
        tc.value = total
    tc.font = BOLD; tc.number_format = fmt; tc.fill = FILL_GREY
    return row + 1


def kv(ws, row, label, value, fmt=None, font=BLUE, note="", yellow=False, col=2):
    c = ws.cell(row, 1, label); c.font = BLACK
    v = ws.cell(row, col, value); v.font = font
    if fmt:
        v.number_format = fmt
    if yellow:
        v.fill = FILL_YEL
    v.border = BOX
    if note:
        n = ws.cell(row, col+1, note); n.font = NOTE
    return row + 1


# ============================================================== 1. README
ws = sheet("README")
title(ws, "GOVHUB — SAAS METRICS & PROJECTIONS WORKBOOK  (v2, rebuilt on real data)",
      "Rebuilt 2026-08-13. Every actual in this file was pulled from a live system and is cited. Nothing is illustrative.", 8)
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 116
r = 4

ws.cell(r, 2, "⚠ WHAT CHANGED FROM v1 — READ THIS FIRST").font = WARN
ws.cell(r, 2).fill = FILL_RED
r += 1
for t in [
  "The July 2026 version of this workbook reported $1,085 MRR, 5 paying customers, 4,250 organic clicks and 10,000 cold",
  "emails sent. None of that was real. It was placeholder data written to demonstrate the formulas, labelled \"ILLUSTRATIVE",
  "SAMPLE DATA\" in the README, and never overwritten. Taking those numbers into an investor conversation would have been",
  "a serious problem. This version replaces every one of them with a queried, sourced actual.",
  "",
  "The honest headline as of 2026-08-13:  $0 MRR · 0 paying customers · 0 cold emails sent · 30 signups (9 external).",
  "GovHub is pre-revenue with a built-but-unlaunched acquisition engine. That is a legitimate stage. Overstating it is not.",
]:
    ws.cell(r, 2, t).font = BLACK if not t.startswith("The honest") else BOLD
    r += 1
r += 1

ws.cell(r, 2, "HOW TO READ THIS WORKBOOK").font = H2; r += 1
for t in [
  "BLUE cells = hardcoded inputs (an actual you pulled, or a lever you set).   BLACK = formulas, do not overwrite.",
  "GREEN = links to another tab.   YELLOW fill = the assumptions worth arguing about — the levers.",
  "Actuals tabs run Sep-25 → Aug-26 (Aug-26 is month-to-date through the 13th). Projections run forward from Sep-26.",
  "Every actual carries its source in column B. If a cell has no source, it is a forward assumption, not a fact.",
]:
    ws.cell(r, 2, t).font = BLACK; r += 1
r += 1

ws.cell(r, 2, "TAB GUIDE").font = H2; r += 1
for name, desc in [
  ("Dashboard",      "Where the business actually is today, what is moving, and what cannot be measured yet (and why)."),
  ("Funnel",         "THE CENTREPIECE. Channel → visitor → signup → activated → engaged → paid, with the conversion rate at every step, per channel."),
  ("Calculator",     "The scenario engine. Change conversion rates, pricing, churn or spend and watch CAC, LTV, LTV:CAC, payback and time-to-$10k MRR move. Sensitivity grids at the bottom."),
  ("Revenue",        "Customer and MRR waterfall. Currently all zeros — the structure is live and will populate on the first Stripe charge."),
  ("SEO",            "Real Semrush + Google Search Console data. Keyword growth is the one genuinely strong number in this file."),
  ("Cold Email",     "Real state of the Instantly build: 16 domains, 12 mailboxes, 3 campaigns, 0 sends. Plus the capacity model for what it produces once launched."),
  ("Expenses",       "The exact cost stack you gave me, plus real Anthropic API spend. Run-rate, annual renewals, and burn."),
  ("Unit Economics", "CAC, LTV, LTV:CAC, payback. Most are undefined until the first paying customer — this tab shows exactly what each one needs."),
  ("Projections",    "24-month driver model (Sep-26 → Aug-28) wired to the Calculator's levers."),
  ("Benchmarks",     "What good looks like, with sources, so every status flag in this file is defensible."),
]:
    ws.cell(r, 2, f"{name} — {desc}").font = BLACK; r += 1
r += 1

ws.cell(r, 2, "DATA SOURCES USED TO BUILD THIS VERSION").font = H2; r += 1
for t in [
  "Supabase project GovProp (iqekrwearenblsmhvdjn) — accounts, subscriptions, plans, user_profiles, v2_submissions,",
  "     v2_token_usage, scheduled_emails. Queried live 2026-08-13.",
  "Semrush us database — govhub.online domain_rank + resource_rank_history + resource_organic. Pulled 2026-08-13.",
  "Google Search Console — via strata-parse/docs/seo/gsc-review-2026-07.md (property verified 2026-06-20).",
  "Instantly — via govhub-marketing/docs/instantly-wave1.md (campaign state as of 2026-08-12).",
  "Resend vendor outreach ledger — govhub-marketing/data/vendor-outreach.json (48 sent, 43 no-contact).",
  "Cost stack — provided directly by the founder, 2026-08-13. Marked as such on the Expenses tab.",
]:
    ws.cell(r, 2, t).font = SMALL; r += 1
r += 1

ws.cell(r, 2, "TWO OPEN ITEMS YOU NEED TO RESOLVE").font = H2; r += 1
for t in [
  "1. Stripe is not wired up. Every row in the plans table has a null stripe_price_id (monthly and annual). There is no",
  "   code path that can take money today. Until that ships, trial → paid conversion is structurally 0%, not behaviourally 0%.",
  "2. Supabase is not in your cost list. The GovProp project is ACTIVE_HEALTHY and serving production. If it is on the free",
  "   tier, it is a capacity risk; if it is billed, it is a missing cost line. Expenses row is stubbed at $0 pending your answer.",
]:
    ws.cell(r, 2, t).font = BLACK; r += 1


# ============================================================== 2. DASHBOARD
ws = sheet("Dashboard")
title(ws, "EXECUTIVE DASHBOARD — as of 2026-08-13",
      "Pre-revenue. This tab separates what is true, what is moving, and what cannot yet be measured.", 7)
for col, w in zip("ABCDEFG", [40, 16, 16, 14, 30, 16, 46]):
    ws.column_dimensions[col].width = w
r = 4

r = band(ws, r, "WHERE THE BUSINESS ACTUALLY IS", 7)
hdr = ["Metric", "Value", "Prior mo", "Δ", "Benchmark / target", "Status", "What it means"]
for i, h in enumerate(hdr):
    c = ws.cell(r, i+1, h); c.font = BOLD; c.fill = FILL_GREY
r += 1

DASH_TOP = r
rows = [
  ("MRR",                    "=Revenue!O24",  "=Revenue!N24", CUR,  "Any positive number",     "PRE-REVENUE",
   "0 Stripe subscriptions exist. Not a churn problem — billing has never been switched on."),
  ("Paying customers",       "=Revenue!O14",  "=Revenue!N14", NUM,  "First 10 = proof",        "PRE-REVENUE",
   "All 30 accounts are plan_type='trial'."),
  ("ARR (run-rate)",         "=B{}*12".format(DASH_TOP), "=C{}*12".format(DASH_TOP), CUR, "—", "PRE-REVENUE", ""),
  ("Total signups (all-time)","=Funnel!D12",  None,           NUM,  "Volume driver",           "REAL",
   "30 accounts created since 2025-09."),
  ("External signups (all-time)","=Funnel!D13",None,          NUM,  "The number that counts",  "REAL",
   "21 of the 30 are founder aliases or test accounts. This is the real top of funnel."),
  ("Active trials right now","=Funnel!D30",   None,           NUM,  "—",                       "REAL",
   "26 trials have already expired without a conversion path existing."),
]
for label, f, pf, fmt, bench, status, meaning in rows:
    ws.cell(r, 1, label).font = BOLD
    c = ws.cell(r, 2, f); c.font = GREEN; c.number_format = fmt
    if pf:
        c = ws.cell(r, 3, pf); c.font = GREEN; c.number_format = fmt
        c = ws.cell(r, 4, f"=IF(OR(B{r}=\"\",C{r}=\"\"),\"\",B{r}-C{r})"); c.font = BLACK; c.number_format = fmt
    ws.cell(r, 5, bench).font = SMALL
    c = ws.cell(r, 6, status); c.font = WARN if status == "PRE-REVENUE" else BOLD
    c.fill = FILL_RED if status == "PRE-REVENUE" else FILL_GRN
    c.alignment = Alignment(horizontal="center")
    ws.cell(r, 7, meaning).font = SMALL
    r += 1
r += 1

r = band(ws, r, "WHAT IS ACTUALLY MOVING — the leading indicators", 7)
for i, h in enumerate(["Metric", "Now", "60 days ago", "Δ", "Why it matters", "Status", "Detail"]):
    c = ws.cell(r, i+1, h); c.font = BOLD; c.fill = FILL_GREY
r += 1
lead = [
  ("Organic keywords ranking (top 100)", "=SEO!O8",  "=SEO!L8",  NUM, "Fastest-moving real asset", "STRONG",
   "1 → 114 in eight weeks. This is the strongest number in the business."),
  ("Est. monthly organic sessions",      "=SEO!O9",  "=SEO!L9",  NUM, "Traffic follows keywords",  "EARLY",
   "Still ~20/mo. Keywords are ranked but deep — position 23 on the big one."),
  ("Proposals generated (v2_submissions)","=Funnel!D22", None,   NUM, "Proof the product runs",   "REAL",
   "30 real proposal runs. The product demonstrably works end to end."),
  ("Anthropic API spend / mo",           "=Expenses!N13","=Expenses!M13",CUR2,"Variable COGS",      "LOW",
   "$26.75 in the heaviest month. COGS is not the constraint."),
  ("Cold emails sent",                   "=('Cold Email'!O8)", None, NUM, "Channel not yet live",  "NOT LIVE",
   "16 domains bought, 12 mailboxes warmed, 3 campaigns drafted, 0 sends."),
]
for label, f, pf, fmt, why, status, detail in lead:
    ws.cell(r, 1, label).font = BOLD
    c = ws.cell(r, 2, f); c.font = GREEN; c.number_format = fmt
    if pf:
        c = ws.cell(r, 3, pf); c.font = GREEN; c.number_format = fmt
        c = ws.cell(r, 4, f"=B{r}-C{r}"); c.font = BLACK; c.number_format = fmt
    ws.cell(r, 5, why).font = SMALL
    c = ws.cell(r, 6, status); c.font = BOLD
    c.fill = {"STRONG": FILL_GRN, "NOT LIVE": FILL_RED}.get(status, FILL_YEL)
    c.alignment = Alignment(horizontal="center")
    ws.cell(r, 7, detail).font = SMALL
    r += 1
r += 1

r = band(ws, r, "WHAT CANNOT BE MEASURED YET — and exactly what each one needs", 7)
for i, h in enumerate(["Metric", "Status", "", "", "Blocked on", "", "First honest read available"]):
    c = ws.cell(r, i+1, h); c.font = BOLD; c.fill = FILL_GREY
r += 1
blocked = [
  ("Trial → paid conversion %", "Stripe billing wired up + first trial cohort passed through it", "~30 days after billing ships"),
  ("Logo churn / gross MRR churn", "At least one paying cohort with 60+ days of history", "~90 days after first charge"),
  ("LTV", "A real churn rate. LTV = ARPA × GM% ÷ churn — churn is the input you do not have", "~90 days after first charge"),
  ("CAC (blended and per channel)", "Paid customers in the denominator. Spend is known; customers are 0", "First 5–10 customers"),
  ("LTV : CAC and CAC payback", "Both of the above", "~120 days after first charge"),
  ("Cold email reply / meeting rates", "Campaigns actually sending. Earliest defensible launch is 2026-08-27", "~2 weeks after launch"),
]
for metric, blocker, when in blocked:
    ws.cell(r, 1, metric).font = BOLD
    c = ws.cell(r, 2, "NOT YET MEASURABLE"); c.font = NOTE
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ws.cell(r, 5, blocker).font = SMALL
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
    ws.cell(r, 7, when).font = SMALL
    r += 1
r += 1

r = band(ws, r, "THE THREE NUMBERS THAT UNLOCK THIS ENTIRE MODEL", 7)
for n, (t, d) in enumerate([
  ("1.  Ship Stripe billing.",
   "Until a price ID exists, trial → paid is 0% by construction. Every unit-economic number in this workbook is blocked behind this one deploy."),
  ("2.  Launch cold email on 2026-08-27.",
   "12 mailboxes × 15–30/day ≈ 5,400 sends/month at full ramp. At benchmark rates that is the first statistically real conversion data you will own."),
  ("3.  Move 'govwin' from position 23 into the top 10.",
   "9,900 monthly searches, already ranked, already 70% of your organic traffic. Position 23 → 8 is roughly a 10–15× click multiple on one page."),
]):
    ws.cell(r, 1, t).font = BOLD
    ws.cell(r, 2, d).font = BLACK
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    r += 1


# ============================================================== 3. FUNNEL
ws = sheet("Funnel")
title(ws, "ACQUISITION FUNNEL — channel → visitor → signup → activated → engaged → paid",
      "Every stage with its conversion rate off the prior stage. All-time actuals, 2025-09-01 → 2026-08-13.", 9)
for col, w in zip("ABCDEFGHI", [38, 34, 14, 14, 14, 14, 16, 18, 52]):
    ws.column_dimensions[col].width = w
r = 4

ws.cell(r, 1, "Read this funnel top to bottom. The conversion column is the rate off the stage directly above it.").font = NOTE
r += 2

r = band(ws, r, "WHOLE-BUSINESS FUNNEL (all channels, all time)", 9)
for i, h in enumerate(["Stage", "Definition", "Count", "Conv. from prior", "Benchmark", "Status", "Source"]):
    c = ws.cell(r, i+1, h); c.font = BOLD; c.fill = FILL_GREY
    if i == 6:
        ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=9)
r += 1
F_TOP = r
stages = [
  ("Reach (impressions + emails sent)", "Everything that could have seen us", None, None, "—",
   "NOT PULLED — export total impressions from Search Console and type them here"),
  ("Visitors (organic sessions)", "Landed on www.govhub.online", 41, "prior", "0.5–2% of impressions",
   "Semrush est. sessions Jul+Aug 2026"),
  ("Signups (all accounts)", "Created an account", 30, "prior", "0.5–2% of visitors",
   "Supabase user_profiles, deleted_at is null"),
  ("— of which external", "Not a founder alias or test account", 9, None, "—",
   "Email-domain classification of the 30"),
  ("Activated", "Completed onboarding", 21, "prior", "40–60% of signups",
   "user_profiles.onboarding_completed = true"),
  ("Engaged", "Ran ≥1 real proposal (v2_submissions)", 30, None, "—",
   "v2_submissions, 2026-05 → 2026-08. Runs, not users."),
  ("Paid", "Has a Stripe subscription", 0, "prior", "10–25% of trials",
   "Supabase subscriptions: 0 rows with a stripe_subscription_id"),
]
for i, (stage, defn, cnt, conv, bench, src) in enumerate(stages):
    ws.cell(r, 1, stage).font = BOLD if not stage.startswith("—") else BLACK
    ws.cell(r, 1).alignment = Alignment(indent=0 if not stage.startswith("—") else 2)
    ws.cell(r, 2, defn).font = SMALL
    c = ws.cell(r, 3, cnt); c.font = BLUE; c.number_format = NUM; c.border = BOX
    if cnt is None:
        c.fill = FILL_YEL
    if conv == "prior":
        prev = r - 1 if not stages[i-1][0].startswith("—") else r - 2
        c = ws.cell(r, 4, f"=IFERROR(C{r}/C{prev},\"\")")
        c.font = BLACK; c.number_format = PCT2
    ws.cell(r, 5, bench).font = SMALL
    ws.cell(r, 7, src).font = SMALL
    ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=9)
    r += 1
F_PAID = r - 1
r += 1

ws.cell(r, 1, "Overall visitor → paid").font = BOLD
c = ws.cell(r, 4, f"=IFERROR(C{F_PAID}/C{F_TOP+1},0)"); c.font = BLACK; c.number_format = PCT2
ws.cell(r, 5, "Undefined in practice — numerator is 0 because billing does not exist yet.").font = NOTE
ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=9)
r += 2

# ---- per channel
r = band(ws, r, "BY CHANNEL — where the 30 signups actually came from", 9)
for i, h in enumerate(["Channel", "Status", "Reach", "Visitors", "Signups", "Activated", "Paid", "Signup→Paid", "Note"]):
    c = ws.cell(r, i+1, h); c.font = BOLD; c.fill = FILL_GREY
r += 1
CH_TOP = r
channels = [
  ("SEO / organic", "LIVE, EARLY", None, 41, 1, 1, 0,
   "signup_source='google' first appears 2026-07. Exactly 1 attributable organic signup."),
  ("Cold email (Instantly)", "NOT LAUNCHED", 0, 0, 0, 0, 0,
   "3 campaigns in Draft, 0 leads, 0 sends. Earliest defensible launch 2026-08-27."),
  ("Vendor / PR outreach (Resend)", "LIVE", 48, 0, 0, 0, 0,
   "48 vendor emails sent Jul–Aug. No reply/click tracking wired up — see gap below."),
  ("Direct / manual / founder-led", "LIVE", 0, 0, 29, 20, 0,
   "signup_source='manual'. Includes all 21 internal/test accounts and 8 external hand-sourced signups."),
]
for name, status, reach, vis, sign, act, paid, note in channels:
    ws.cell(r, 1, name).font = BOLD
    c = ws.cell(r, 2, status); c.font = BOLD
    c.fill = {"NOT LAUNCHED": FILL_RED, "LIVE, EARLY": FILL_YEL}.get(status, FILL_GRN)
    c.alignment = Alignment(horizontal="center")
    for j, v in enumerate([reach, vis, sign, act, paid]):
        cc = ws.cell(r, 3+j, v); cc.font = BLUE; cc.number_format = NUM; cc.border = BOX
    c = ws.cell(r, 8, f"=IFERROR(G{r}/E{r},0)"); c.font = BLACK; c.number_format = PCT2
    ws.cell(r, 9, note).font = SMALL
    r += 1
CH_BOT = r - 1
ws.cell(r, 1, "TOTAL").font = BOLD
for j, col in enumerate("CDEFG"):
    c = ws.cell(r, 3+j, f"=SUM({col}{CH_TOP}:{col}{CH_BOT})")
    c.font = BOLD; c.number_format = NUM; c.fill = FILL_GREY
c = ws.cell(r, 8, f"=IFERROR(G{r}/E{r},0)"); c.font = BOLD; c.number_format = PCT2; c.fill = FILL_GREY
r += 2

# ---- trial mechanics
r = band(ws, r, "TRIAL MECHANICS — what happened to the 30 trials", 9)
for i, h in enumerate(["Metric", "Count", "% of trials", "", "Note"]):
    c = ws.cell(r, i+1, h); c.font = BOLD; c.fill = FILL_GREY
r += 1
T_TOP = r
trials = [
  ("Trials started (all-time)", 29, "First 2025-10-19, most recent 2026-08-11. accounts.trial_start"),
  ("Trials expired", 26, "trial_end <= now(). Every one of these expired with no way to pay."),
  ("Trials active right now", 3, "trial_end > now()"),
  ("Trials converted to paid", 0, "Zero. There is no Stripe price ID to convert against."),
]
for name, cnt, note in trials:
    ws.cell(r, 1, name).font = BOLD
    c = ws.cell(r, 2, cnt); c.font = BLUE; c.number_format = NUM; c.border = BOX
    c = ws.cell(r, 3, f"=IFERROR(B{r}/$B${T_TOP},0)"); c.font = BLACK; c.number_format = PCT
    ws.cell(r, 5, note).font = SMALL
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=9)
    r += 1
r += 1

# ---- monthly trend
r = band(ws, r, "MONTHLY TREND — signups, activation and product usage", 9)
r2 = r
ws2_start = r
r = monthhdr(ws, r, "Metric", "Source")
r = datarow(ws, r, "Signups (all)", "Supabase user_profiles", SIGNUPS)
r = datarow(ws, r, "Signups — external", "Non-founder / non-test email", SIGNUPS_EXT)
r = datarow(ws, r, "Signups — internal / test", "Founder aliases + QA accounts", SIGNUPS_INT)
row_onb = r
r = datarow(ws, r, "Activated (onboarding complete)", "user_profiles.onboarding_completed", ONBOARDED)
r = formularow(ws, r, "Activation rate", "Activated ÷ all signups",
               "=IFERROR({c}%d/{c}%d,\"\")" % (row_onb, row_onb-3), PCT,
               total=f"=IFERROR({TOTC}{row_onb}/{TOTC}{row_onb-3},\"\")")
r = datarow(ws, r, "Proposals generated", "v2_submissions (runs, not users)", SUBMISSIONS)
r = datarow(ws, r, "Lifecycle emails sent", "scheduled_emails status='sent'", LIFECYCLE_EM)
r = datarow(ws, r, "Paying customers (ending)", "Supabase subscriptions w/ Stripe id", [0]*12, NUM, "last")

ws.column_dimensions['A'].width = 38
ws.column_dimensions['B'].width = 34


# ============================================================== 4. CALCULATOR
ws = sheet("Calculator")
title(ws, "SCENARIO CALCULATOR — conversion rates, pricing, CAC, CLTV",
      "Change any BLUE cell. Everything black recalculates. Yellow = the levers worth arguing about with an investor.", 10)
for col, w in zip("ABCDEFGHIJ", [42, 15, 15, 15, 15, 15, 15, 6, 44, 20]):
    ws.column_dimensions[col].width = w
r = 4

# --- pricing
r = band(ws, r, "1 · PRICING & MIX", 10)
P_TOP = r
r = kv(ws, r, "Solo price ($/mo)", 129, CUR, BLUE, "Live price in plans table", True)
r = kv(ws, r, "Pro price ($/mo)", 349, CUR, BLUE, "Live price in plans table", True)
r = kv(ws, r, "Team price ($/mo)", 699, CUR, BLUE, "Live price in plans table", True)
P_SOLO, P_PRO, P_TEAM = P_TOP, P_TOP+1, P_TOP+2
r = kv(ws, r, "Mix — Solo %", 0.60, PCT, BLUE, "Assumption: SMB-heavy govcon ICP", True)
r = kv(ws, r, "Mix — Pro %", 0.32, PCT, BLUE, "", True)
r = kv(ws, r, "Mix — Team %", 0.08, PCT, BLUE, "", True)
M_SOLO, M_PRO, M_TEAM = P_TOP+3, P_TOP+4, P_TOP+5
c = ws.cell(r, 1, "Mix check (must = 100%)"); c.font = BLACK
c = ws.cell(r, 2, f"=B{M_SOLO}+B{M_PRO}+B{M_TEAM}"); c.font = BLACK; c.number_format = PCT
ws.cell(r, 3, "=IF(ROUND(B{},4)=1,\"OK\",\"FIX MIX\")".format(r)).font = BOLD
r += 1
ARPA = r
c = ws.cell(r, 1, "Blended ARPA ($/mo)"); c.font = BOLD
c = ws.cell(r, 2, f"=B{P_SOLO}*B{M_SOLO}+B{P_PRO}*B{M_PRO}+B{P_TEAM}*B{M_TEAM}")
c.font = BLACK; c.number_format = CUR; c.fill = FILL_GRN; c.border = BOX
ws.cell(r, 3, "Weighted average revenue per account").font = NOTE
r += 1
r = kv(ws, r, "Annual-plan share of new customers %", 0.20, PCT, BLUE,
       "Annual = 10× monthly in the plans table (2 months free)", True)
ANN_SHARE = r - 1
r = kv(ws, r, "Annual discount (implied)", 0.1667, PCT, BLUE, "1 - 10/12", False)
ANN_DISC = r - 1
EFF_ARPA = r
c = ws.cell(r, 1, "Effective ARPA after annual discount"); c.font = BOLD
c = ws.cell(r, 2, f"=B{ARPA}*(1-B{ANN_SHARE}*B{ANN_DISC})")
c.font = BLACK; c.number_format = CUR; c.fill = FILL_GRN; c.border = BOX
ws.cell(r, 3, "This is the ARPA used everywhere below").font = NOTE
r += 2

# --- SEO funnel
r = band(ws, r, "2 · FUNNEL — SEO / ORGANIC", 10)
S_TOP = r
r = kv(ws, r, "Monthly organic sessions", 41, NUM, BLUE, "Actual Jul+Aug 2026 (Semrush est.)", True)
r = kv(ws, r, "Session → signup %", 0.025, PCT, BLUE, "Benchmark 0.5–2%; you are early and low-traffic", True)
r = kv(ws, r, "Signup → activated %", 0.70, PCT, BLUE, "Actual all-time: 70% (21/30)", True)
r = kv(ws, r, "Activated → paid %", 0.12, PCT, BLUE, "ASSUMPTION — no actual exists. Benchmark 10–25%", True)
S_SESS, S_S2S, S_S2A, S_A2P = S_TOP, S_TOP+1, S_TOP+2, S_TOP+3
S_SIGN = r
c = ws.cell(r, 1, "→ Signups / mo"); c.font = BLACK; c.alignment = Alignment(indent=1)
c = ws.cell(r, 2, f"=B{S_SESS}*B{S_S2S}"); c.font = BLACK; c.number_format = NUM1
r += 1
S_ACT = r
c = ws.cell(r, 1, "→ Activated / mo"); c.font = BLACK; c.alignment = Alignment(indent=1)
c = ws.cell(r, 2, f"=B{S_SIGN}*B{S_S2A}"); c.font = BLACK; c.number_format = NUM1
r += 1
S_CUST = r
c = ws.cell(r, 1, "→ New customers / mo (SEO)"); c.font = BOLD; c.alignment = Alignment(indent=1)
c = ws.cell(r, 2, f"=B{S_ACT}*B{S_A2P}"); c.font = BLACK; c.number_format = NUM1
c.fill = FILL_GRN; c.border = BOX
r += 2

# --- cold email funnel
r = band(ws, r, "3 · FUNNEL — COLD EMAIL", 10)
E_TOP = r
r = kv(ws, r, "Mailboxes sending", 12, NUM, BLUE, "Wave 1: 12 mailboxes across 12 domains", True)
r = kv(ws, r, "Sends per mailbox per day", 15, NUM, BLUE, "Instantly daily cap; 50–100 only after months of history", True)
r = kv(ws, r, "Sending days per month", 22, NUM, BLUE, "Weekdays only", True)
E_MB, E_PER, E_DAYS = E_TOP, E_TOP+1, E_TOP+2
E_SENT = r
c = ws.cell(r, 1, "→ Emails sent / mo"); c.font = BLACK; c.alignment = Alignment(indent=1)
c = ws.cell(r, 2, f"=B{E_MB}*B{E_PER}*B{E_DAYS}"); c.font = BLACK; c.number_format = NUM
r += 1
r = kv(ws, r, "Bounce rate %", 0.02, PCT, BLUE, "Keep <2% — MillionVerifier is the control here", True)
E_BOUNCE = r - 1
E_DELIV = r
c = ws.cell(r, 1, "→ Delivered / mo"); c.font = BLACK; c.alignment = Alignment(indent=1)
c = ws.cell(r, 2, f"=B{E_SENT}*(1-B{E_BOUNCE})"); c.font = BLACK; c.number_format = NUM
r += 1
r = kv(ws, r, "Reply rate % (of delivered)", 0.034, PCT, BLUE, "2026 industry avg ~3.4%", True)
r = kv(ws, r, "Positive share of replies %", 0.33, PCT, BLUE, "Rest are 'remove me' / not-now", True)
r = kv(ws, r, "Positive reply → meeting %", 0.42, PCT, BLUE, "", True)
r = kv(ws, r, "Meeting → trial %", 0.65, PCT, BLUE, "", True)
r = kv(ws, r, "Trial → paid % (email)", 0.30, PCT, BLUE, "Higher intent than organic — they spoke to you", True)
E_RR, E_POS, E_MTG, E_TRIAL, E_PAID = E_DELIV+1, E_DELIV+2, E_DELIV+3, E_DELIV+4, E_DELIV+5
E_NREP = r
c = ws.cell(r, 1, "→ Replies / mo"); c.font = BLACK; c.alignment = Alignment(indent=1)
c = ws.cell(r, 2, f"=B{E_DELIV}*B{E_RR}"); c.font = BLACK; c.number_format = NUM1
r += 1
E_NPOS = r
c = ws.cell(r, 1, "→ Positive replies / mo"); c.font = BLACK; c.alignment = Alignment(indent=1)
c = ws.cell(r, 2, f"=B{E_NREP}*B{E_POS}"); c.font = BLACK; c.number_format = NUM1
r += 1
E_NMTG = r
c = ws.cell(r, 1, "→ Meetings / mo"); c.font = BLACK; c.alignment = Alignment(indent=1)
c = ws.cell(r, 2, f"=B{E_NPOS}*B{E_MTG}"); c.font = BLACK; c.number_format = NUM1
r += 1
E_NTRI = r
c = ws.cell(r, 1, "→ Trials / mo"); c.font = BLACK; c.alignment = Alignment(indent=1)
c = ws.cell(r, 2, f"=B{E_NMTG}*B{E_TRIAL}"); c.font = BLACK; c.number_format = NUM1
r += 1
E_CUST = r
c = ws.cell(r, 1, "→ New customers / mo (email)"); c.font = BOLD; c.alignment = Alignment(indent=1)
c = ws.cell(r, 2, f"=B{E_NTRI}*B{E_PAID}"); c.font = BLACK; c.number_format = NUM1
c.fill = FILL_GRN; c.border = BOX
r += 2

# --- other
r = band(ws, r, "4 · FUNNEL — DIRECT / REFERRAL / FOUNDER-LED", 10)
O_TOP = r
r = kv(ws, r, "Direct signups / mo", 2, NUM, BLUE, "Hand-sourced, demos, network", True)
r = kv(ws, r, "Trial → paid % (direct)", 0.25, PCT, BLUE, "Warmest channel", True)
O_CUST = r
c = ws.cell(r, 1, "→ New customers / mo (direct)"); c.font = BOLD; c.alignment = Alignment(indent=1)
c = ws.cell(r, 2, f"=B{O_TOP}*B{O_TOP+1}"); c.font = BLACK; c.number_format = NUM1
c.fill = FILL_GRN; c.border = BOX
r += 2

# --- retention & spend
r = band(ws, r, "5 · RETENTION, MARGIN & SPEND", 10)
R_TOP = r
r = kv(ws, r, "Monthly logo churn %", 0.05, PCT, BLUE, "SMB SaaS 3–7%. THE most sensitive lever in the model", True)
r = kv(ws, r, "Monthly expansion %", 0.01, PCT, BLUE, "Solo→Pro upgrades, Team seats", True)
r = kv(ws, r, "Gross margin %", 0.85, PCT, BLUE, "API cost is tiny — $26.75 in your heaviest month", True)
CHURN, EXPAN, GM = R_TOP, R_TOP+1, R_TOP+2
r = kv(ws, r, "SEO spend / mo (cash)", 0, CUR, BLUE, "Content is produced in-house — no cash cost today", True)
r = kv(ws, r, "Cold email spend / mo (cash)", 0, CUR, GREEN, "Live link → Expenses marketing subtotal")
r = kv(ws, r, "Other S&M spend / mo (cash)", 0, CUR, BLUE, "Paid ads, events", True)
SP_SEO, SP_EMAIL, SP_OTHER = R_TOP+3, R_TOP+4, R_TOP+5
r = kv(ws, r, "Founder / labour cost allocated to S&M", 5000, CUR, BLUE,
       "⚠ THE ASSUMPTION THAT MAKES CAC CREDIBLE — see note below", True)
SP_LABOUR = r - 1
ws.cell(r, 1, "⚠  WHY THIS CELL IS NOT ZERO").font = WARN
ws.cell(r, 1).fill = FILL_RED
ws.cell(r, 2, "Default is $5,000/mo ≈ 50% of a $120k/yr founder comp, charged to sales & marketing. "
              "Set it to $0 and CAC collapses to the software bill alone, producing an LTV:CAC around 65x — "
              "a number no investor will believe and you should never present. Imputing your own time is what "
              "makes the CAC on this tab defensible in a diligence conversation.").font = WARN
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
r += 2

# --- outputs
r = band(ws, r, "6 · OUTPUTS — the numbers an investor will ask for", 10)
OUT_TOP = r
for i, h in enumerate(["Metric", "Value", "", "Benchmark", "", "", "", "", "How it is calculated"]):
    c = ws.cell(r, i+1, h); c.font = BOLD; c.fill = FILL_GREY
r += 1
TOT_CUST = r
outs = [
  ("New customers / mo (all channels)", f"=B{S_CUST}+B{E_CUST}+B{O_CUST}", NUM1, "—",
   "SEO + cold email + direct"),
  ("New MRR added / mo", f"=B{TOT_CUST}*B{EFF_ARPA}", CUR, "—",
   "New customers × effective ARPA"),
  ("Total S&M spend / mo (fully loaded)", f"=B{SP_SEO}+B{SP_EMAIL}+B{SP_OTHER}+B{SP_LABOUR}", CUR, "—",
   "SEO + email + other + founder/labour time"),
  ("Blended CAC", f"=IFERROR(B{TOT_CUST+2}/B{TOT_CUST},0)", CUR, "Judge only against LTV",
   "Total S&M ÷ new customers"),
  ("SEO CAC (cash only)", f"=IFERROR(B{SP_SEO}/B{S_CUST},0)", CUR, "—",
   "SEO cash spend ÷ SEO customers. Excludes your time."),
  ("Cold email CAC (cash only)", f"=IFERROR(B{SP_EMAIL}/B{E_CUST},0)", CUR, "—",
   "Email cash spend ÷ email customers. Excludes your time."),
  ("LTV (gross-margin adjusted)", f"=IFERROR(B{EFF_ARPA}*B{GM}/B{CHURN},0)", CUR, "—",
   "ARPA × GM% ÷ monthly churn  (Skok)"),
  ("LTV : CAC", f"=IFERROR(B{TOT_CUST+6}/B{TOT_CUST+3},0)", MULT, "≥3x healthy, 5x+ strong",
   "LTV ÷ blended CAC"),
  ("CAC payback (months)", f"=IFERROR(B{TOT_CUST+3}/(B{EFF_ARPA}*B{GM}),0)", NUM1, "≤12 months",
   "CAC ÷ (ARPA × GM%)"),
  ("Steady-state MRR ceiling", f"=IFERROR(B{TOT_CUST+1}/B{CHURN},0)", CUR, "—",
   "New MRR ÷ churn. The MRR this engine converges to and CANNOT exceed."),
  ("Months to $10k MRR", None, NUM1, "—",
   "Solved from the churn-adjusted MRR curve. 'never' means the ceiling above is below $10k."),
  ("Months to cover total cost", None, NUM1, "—",
   "Month at which gross profit covers Expenses cash burn PLUS the founder/labour cost above."),
]
for label, f, fmt, bench, how in outs:
    ws.cell(r, 1, label).font = BOLD
    if f:
        c = ws.cell(r, 2, f); c.font = BLACK; c.number_format = fmt
        c.fill = FILL_GRN; c.border = BOX
    ws.cell(r, 4, bench).font = SMALL
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
    ws.cell(r, 9, how).font = SMALL
    r += 1
M10K = TOT_CUST + 10
MBRK = TOT_CUST + 11
CEIL = TOT_CUST + 9
NEWMRR = TOT_CUST + 1
LTV = TOT_CUST + 6
BCAC = TOT_CUST + 3

c = ws.cell(M10K, 2, f'=IF(B{CEIL}<=10000,"never",LN(1-10000*B{CHURN}/B{NEWMRR})/LN(1-B{CHURN}))')
c.font = BLACK; c.number_format = NUM1; c.fill = FILL_GRN; c.border = BOX
c = ws.cell(MBRK, 2,
    f'=IF(B{CEIL}*B{GM}<=Expenses!$C$34,"never",'
    f'LN(1-(Expenses!$C$34/B{GM})*B{CHURN}/B{NEWMRR})/LN(1-B{CHURN}))')
c.font = BLACK; c.number_format = NUM1; c.fill = FILL_GRN; c.border = BOX
r += 1

# --- sensitivity grids
r = band(ws, r, "7 · SENSITIVITY — where the model actually breaks", 10)
r += 1

# Grid A: LTV = ARPA x GM / churn
ws.cell(r, 1, "A ·  LTV  —  rows = monthly churn, columns = blended ARPA").font = BOLD
r += 1
GA = r
ARPAS = [150, 200, 250, 300, 350]
CHURNS = [0.03, 0.04, 0.05, 0.07, 0.10]
ws.cell(r, 1, "monthly churn ↓ / ARPA →").font = NOTE
for j, a in enumerate(ARPAS):
    c = ws.cell(r, 2+j, a); c.font = BOLD; c.number_format = CUR
    c.fill = FILL_GREY; c.alignment = Alignment(horizontal="center")
r += 1
for ch in CHURNS:
    c = ws.cell(r, 1, ch); c.font = BOLD; c.number_format = PCT; c.fill = FILL_GREY
    for j in range(len(ARPAS)):
        col = get_column_letter(2+j)
        cc = ws.cell(r, 2+j, f"={col}${GA}*$B${GM}/$A{r}")
        cc.font = BLACK; cc.number_format = CUR; cc.border = BOX
    r += 1
r += 1

# Grid B: LTV:CAC
ws.cell(r, 1, "B ·  LTV : CAC  —  rows = monthly churn, columns = blended ARPA (at current blended CAC)").font = BOLD
r += 1
GB = r
ws.cell(r, 1, "monthly churn ↓ / ARPA →").font = NOTE
for j, a in enumerate(ARPAS):
    c = ws.cell(r, 2+j, a); c.font = BOLD; c.number_format = CUR
    c.fill = FILL_GREY; c.alignment = Alignment(horizontal="center")
r += 1
GB_TOP = r
for ch in CHURNS:
    c = ws.cell(r, 1, ch); c.font = BOLD; c.number_format = PCT; c.fill = FILL_GREY
    for j in range(len(ARPAS)):
        col = get_column_letter(2+j)
        cc = ws.cell(r, 2+j, f"=IFERROR({col}${GB}*$B${GM}/$A{r}/$B${BCAC},0)")
        cc.font = BLACK; cc.number_format = MULT; cc.border = BOX
    r += 1
r += 1

# Grid C: new customers/mo
ws.cell(r, 1, "C ·  NEW CUSTOMERS / MONTH  —  rows = trial→paid %, columns = qualified trials / month").font = BOLD
r += 1
GC = r
TRIALS = [5, 10, 20, 40, 80]
T2P = [0.05, 0.10, 0.15, 0.20, 0.25]
ws.cell(r, 1, "trial→paid ↓ / trials →").font = NOTE
for j, t in enumerate(TRIALS):
    c = ws.cell(r, 2+j, t); c.font = BOLD; c.number_format = NUM
    c.fill = FILL_GREY; c.alignment = Alignment(horizontal="center")
r += 1
for t2p in T2P:
    c = ws.cell(r, 1, t2p); c.font = BOLD; c.number_format = PCT; c.fill = FILL_GREY
    for j in range(len(TRIALS)):
        col = get_column_letter(2+j)
        cc = ws.cell(r, 2+j, f"={col}${GC}*$A{r}")
        cc.font = BLACK; cc.number_format = NUM1; cc.border = BOX
    r += 1
r += 1

# Grid D: months to 10k MRR
ws.cell(r, 1, "D ·  MONTHS TO $10,000 MRR  —  rows = new customers/mo, columns = blended ARPA (at current churn)").font = BOLD
r += 1
GD = r
NEWC = [2, 4, 6, 10, 15]
ws.cell(r, 1, "new cust/mo ↓ / ARPA →").font = NOTE
for j, a in enumerate(ARPAS):
    c = ws.cell(r, 2+j, a); c.font = BOLD; c.number_format = CUR
    c.fill = FILL_GREY; c.alignment = Alignment(horizontal="center")
r += 1
for nc in NEWC:
    c = ws.cell(r, 1, nc); c.font = BOLD; c.number_format = NUM; c.fill = FILL_GREY
    for j in range(len(ARPAS)):
        col = get_column_letter(2+j)
        cc = ws.cell(r, 2+j,
            f'=IF($A{r}*{col}${GD}/$B${CHURN}<=10000,"never",'
            f'LN(1-10000*$B${CHURN}/($A{r}*{col}${GD}))/LN(1-$B${CHURN}))')
        cc.font = BLACK; cc.number_format = NUM1; cc.border = BOX
    r += 1
r += 1
ws.cell(r, 1, '"never" = at that combination the steady-state ceiling (new MRR ÷ churn) is below $10k. '
              'Churn, not acquisition, is what caps you.').font = NOTE


# ============================================================== 5. REVENUE
ws = sheet("Revenue")
title(ws, "REVENUE & CUSTOMERS — actuals",
      "All zeros, and that is the correct number. 0 rows in subscriptions carry a stripe_subscription_id.", 16)
r = 4
ws.cell(r, 1, "Prices below are live values from the Supabase plans table, read 2026-08-13.").font = NOTE
r += 2
r = band(ws, r, "PRICE LIST (source: plans table)", 16)
for name, mo, yr, prop in [("Trial", 0, 0, 3), ("Solo", 129, 1290, 3), ("Pro", 349, 3490, 15), ("Team", 699, 6990, 50)]:
    ws.cell(r, 1, f"{name}").font = BOLD
    c = ws.cell(r, 2, f"${mo}/mo · ${yr}/yr · {prop} proposals/mo"); c.font = BLACK
    ws.cell(r, 3, "stripe_price_id: NULL — cannot be charged").font = WARN if name != "Trial" else NOTE
    r += 1
r += 1

r = band(ws, r, "CUSTOMER MOVEMENT", 16)
r = monthhdr(ws, r, "Metric", "Source")
R_NEW = r
r = datarow(ws, r, "New customers", "Supabase subscriptions", [0]*12)
r = datarow(ws, r, "Churned customers", "Supabase subscriptions", [0]*12)
r = datarow(ws, r, "Ending customers", "Cumulative", [0]*12, NUM, "last")
r += 1
r = band(ws, r, "MRR WATERFALL", 16)
R_BEG = r
r = datarow(ws, r, "Beginning MRR", "", [0]*12, CUR, "last")
r = datarow(ws, r, "New business MRR", "Stripe", [0]*12, CUR)
r = datarow(ws, r, "Expansion MRR", "Stripe", [0]*12, CUR)
r = datarow(ws, r, "Contraction MRR", "Stripe", [0]*12, CUR)
r = datarow(ws, r, "Churned MRR", "Stripe", [0]*12, CUR)
r = datarow(ws, r, "Net new MRR", "", [0]*12, CUR)
R_END = r
r = datarow(ws, r, "Ending MRR", "", [0]*12, CUR, "last")
r = datarow(ws, r, "ARR (run-rate)", "Ending MRR × 12", [0]*12, CUR, "last")
r += 1
ws.cell(r, 1, "When the first charge lands, overwrite these rows from the Stripe dashboard. Every ratio in this "
              "workbook that currently reads 0 or 'never' will start returning a real number.").font = NOTE


# ============================================================== 6. SEO
ws = sheet("SEO")
title(ws, "MARKETING — SEO / ORGANIC  (real data)",
      "Sources: Semrush us database (pulled 2026-08-13) and Google Search Console (property verified 2026-06-20).", 16)
r = 4
r = band(ws, r, "VISIBILITY", 16)
r = monthhdr(ws, r, "Metric", "Source")
SEO_KW_ROW = r
r = datarow(ws, r, "Organic keywords (top 100)", "Semrush resource_rank_history", SEO_KW, NUM, "last")
r = datarow(ws, r, "Est. organic sessions / mo", "Semrush organic traffic estimate", SEO_TRAFFIC, NUM, "last")
r = datarow(ws, r, "GSC clicks (28-day trailing)", "GSC milestone notifications — fill monthly", GSC_CLICKS, NUM, "last")
r = datarow(ws, r, "Est. traffic value ($/mo)", "Semrush organic cost", [0]*10+[94, 241], CUR, "last")
r += 1
ws.cell(r, 1, "Semrush shows 1 keyword and 0 traffic every month from Dec-25 to Jun-26. The site went from 1 → 90 keywords "
              "in July and 90 → 114 in August. That inflection is real and it is the best number in this workbook.").font = NOTE
r += 2

r = band(ws, r, "TOP ORGANIC KEYWORDS (Semrush, 2026-08-13)", 16)
for i, h in enumerate(["Keyword", "Position", "Volume", "CPC", "% of traffic", "Landing page"]):
    c = ws.cell(r, i+1, h); c.font = BOLD; c.fill = FILL_GREY
r += 1
kws = [
  ("govwin", 23, 9900, 12.59, 0.70, "/vs/govwin-iq/"),
  ("gov hub", 7, 140, 10.48, 0.05, "/"),
  ("govwin iq pricing", 14, 260, 14.40, 0.05, "/vs/govwin-iq/"),
  ("responsive io alternative", 29, 140, 0.00, 0.00, "/alternatives/responsive/"),
  ("what does idiq stand for in government contracting", 19, 70, 0.00, 0.00, "/glossary/idiq/"),
  ("cage number sam gov", 31, 40, 2.29, 0.00, "/glossary/cage-code/"),
  ("govtribe pricing", 40, 90, 5.57, 0.00, "/pricing/"),
  ("govwin subscription cost", 58, 110, 7.67, 0.00, "/pricing/"),
  ("rfp", 70, 40500, 3.99, 0.00, "/glossary/rfp/"),
  ("how to write a government proposal", 71, 110, 4.45, 0.00, "/blog/how-to-write-a-government-proposal/"),
]
for k, pos, vol, cpc, share, url in kws:
    ws.cell(r, 1, k).font = BLACK
    c = ws.cell(r, 2, pos); c.font = BLUE; c.number_format = NUM
    c = ws.cell(r, 3, vol); c.font = BLUE; c.number_format = NUM
    c = ws.cell(r, 4, cpc); c.font = BLUE; c.number_format = CUR2
    c = ws.cell(r, 5, share); c.font = BLUE; c.number_format = PCT
    ws.cell(r, 6, url).font = SMALL
    r += 1
r += 1
ws.cell(r, 1, "THE SINGLE HIGHEST-LEVERAGE SEO FACT IN THIS FILE").font = WARN
r += 1
ws.cell(r, 1, "'govwin' — 9,900 searches/month, $12.59 CPC — is ranked at position 23 and already produces 70% of your "
              "organic traffic from page 3. Position 23 → 8 is roughly a 10–15× click multiple on that one page. "
              "Nothing else on this tab is worth doing first.").font = BLACK
r += 2

r = band(ws, r, "ORGANIC FUNNEL & COST", 16)
r = monthhdr(ws, r, "Metric", "Source / note")
SEO_SIGN = r
r = datarow(ws, r, "Organic signups", "user_profiles.signup_source='google'", [0]*10+[1, 0])
r = datarow(ws, r, "Organic new customers", "Supabase subscriptions", [0]*12)
r = datarow(ws, r, "SEO spend ($)", "Content produced in-house — no cash cost booked", [0]*12, CUR)
r = formularow(ws, r, "Session → signup %", "Signups ÷ est. sessions",
               "=IFERROR({c}%d/{c}%d,\"\")" % (SEO_SIGN, SEO_KW_ROW+1), PCT2,
               total=f"=IFERROR({TOTC}{SEO_SIGN}/{TOTC}{SEO_KW_ROW+1},\"\")")
ws.column_dimensions['A'].width = 44
ws.column_dimensions['B'].width = 42


# ============================================================== 7. COLD EMAIL
ws = sheet("Cold Email")
title(ws, "MARKETING — COLD EMAIL  (built, not launched)",
      "Source: govhub-marketing/docs/instantly-wave1.md, campaign state 2026-08-12. Zero emails have been sent.", 16)
r = 4
ws.cell(r, 1, "STATUS: 3 campaigns created 2026-08-12, all in Draft with zero leads. Nothing sends until leads are "
              "uploaded and campaigns started.").font = WARN
ws.cell(r, 1).fill = FILL_RED
r += 2

r = band(ws, r, "ACTUALS", 16)
r = monthhdr(ws, r, "Metric", "Source")
CE_SENT_ROW = r
r = datarow(ws, r, "Cold emails sent", "Instantly — campaigns in Draft", EMAILS_SENT)
r = datarow(ws, r, "Replies", "n/a — nothing sent", [0]*12)
r = datarow(ws, r, "Meetings booked", "n/a — nothing sent", [0]*12)
r = datarow(ws, r, "New customers from cold email", "n/a — nothing sent", [0]*12)
r = datarow(ws, r, "Sending domains owned", "16 purchased 2026-08-06", CE_DOMAINS, NUM, "last")
r = datarow(ws, r, "Mailboxes warmed (wave 1)", "12 live, warmup score 100", CE_MAILBOXES, NUM, "last")
r += 1
r = band(ws, r, "SEPARATE MOTION — vendor / PR outreach via Resend (this one IS sending)", 16)
r = monthhdr(ws, r, "Metric", "Source")
r = datarow(ws, r, "Vendor outreach emails sent", "govhub-marketing/data/vendor-outreach.json", VENDOR_SENT)
r += 1
ws.cell(r, 1, "48 vendor emails have gone out (43 of 91 ledger rows have no contact found). There is no open, click or "
              "reply tracking on this motion — so its contribution to the funnel is currently unmeasurable. Wiring "
              "Resend webhooks into the email_events table (0 rows today) would fix that cheaply.").font = NOTE
r += 2

r = band(ws, r, "LAUNCH READINESS", 16)
for i, h in enumerate(["Item", "State", "Note"]):
    c = ws.cell(r, i+1, h); c.font = BOLD; c.fill = FILL_GREY
r += 1
ready = [
  ("Domains purchased", "16 (2026-08-06)", "$141 first year, $10/yr each on renewal"),
  ("Mailboxes warmed", "12, warmup score 100", "~54 warmup emails each — that is not a sending history"),
  ("Campaigns built", "3 (A serial_bidder, B new_prime, C registered_no_awards)", "All Draft, 0 leads"),
  ("Leads uploaded", "0", "THE BLOCKER — nothing sends without these"),
  ("Earliest defensible launch", "2026-08-27", "Day 21 from domain creation. Day 30 (2026-09-05) is safer"),
  ("Schedule backstop", "start_date 2026-08-27", "An accidental activation still sends nothing before this date"),
]
for item, state, note in ready:
    ws.cell(r, 1, item).font = BOLD
    c = ws.cell(r, 2, state); c.font = BLUE
    ws.cell(r, 3, note).font = SMALL
    r += 1
r += 1

r = band(ws, r, "CAPACITY MODEL — what this engine produces once it is running", 16)
ws.cell(r, 1, "Live from the Calculator tab. Change the levers there, not here.").font = NOTE
r += 1
for i, h in enumerate(["Stage", "Per month", "Rate used"]):
    c = ws.cell(r, i+1, h); c.font = BOLD; c.fill = FILL_GREY
r += 1
cap = [
  ("Emails sent", f"=Calculator!B{E_SENT}", f"=Calculator!B{E_MB}&\" mailboxes × \"&Calculator!B{E_PER}&\"/day × \"&Calculator!B{E_DAYS}&\" days\""),
  ("Delivered", f"=Calculator!B{E_DELIV}", f"=TEXT(1-Calculator!B{E_BOUNCE},\"0.0%\")&\" deliverability\""),
  ("Replies", f"=Calculator!B{E_NREP}", f"=TEXT(Calculator!B{E_RR},\"0.0%\")&\" reply rate\""),
  ("Positive replies", f"=Calculator!B{E_NPOS}", f"=TEXT(Calculator!B{E_POS},\"0.0%\")&\" of replies positive\""),
  ("Meetings", f"=Calculator!B{E_NMTG}", f"=TEXT(Calculator!B{E_MTG},\"0.0%\")&\" positive → meeting\""),
  ("Trials", f"=Calculator!B{E_NTRI}", f"=TEXT(Calculator!B{E_TRIAL},\"0.0%\")&\" meeting → trial\""),
  ("New customers", f"=Calculator!B{E_CUST}", f"=TEXT(Calculator!B{E_PAID},\"0.0%\")&\" trial → paid\""),
]
for stage, f, rate in cap:
    ws.cell(r, 1, stage).font = BOLD
    c = ws.cell(r, 2, f); c.font = GREEN; c.number_format = NUM1; c.border = BOX
    c = ws.cell(r, 3, rate); c.font = SMALL
    r += 1
ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 60


# ============================================================== 8. EXPENSES
ws = sheet("Expenses")
title(ws, "EXPENSES — cost stack, run-rate and burn",
      "Cost stack provided by the founder 2026-08-13. Anthropic API is real spend from v2_token_usage.", 16)
for col, w in zip("ABCDEFG", [34, 26, 14, 14, 16, 18, 46]):
    ws.column_dimensions[col].width = w
r = 4

r = band(ws, r, "THE COST STACK — current run-rate", 7)
for i, h in enumerate(["Vendor / item", "Billing", "$ / month", "$ / year", "Category", "Status", "Note"]):
    c = ws.cell(r, i+1, h); c.font = BOLD; c.fill = FILL_GREY
r += 1

STACK_TOP = r
ops = [
  ("Claude (subscription)", "$100.00 / month", 100.00, "Operational", "Fixed", "Founder-provided"),
  ("Anthropic API", "usage-based", None, "Operational", "Variable", "Actual: $0.19 Jun · $26.75 Jul · $7.90 Aug MTD. Cell links to the trailing month."),
  ("OpenAI / ChatGPT API", "usage-based", 0.00, "Operational", "Variable", "Founder-provided as variable. No spend figure supplied — enter actuals here."),
  ("Resend", "$20.00 / month", 20.00, "Operational", "Fixed", "Founder-provided"),
  ("Cloudflare", "$5.00 / month", 5.00, "Operational", "Fixed", "Founder-provided"),
  ("govhub.online (main domain)", "$40.00 / year", None, "Operational", "Annual", "Renews 9/26 every year. Monthly figure is the $40 amortised."),
  ("Supabase", "unknown", 0.00, "Operational", "OPEN ITEM", "NOT in the founder's cost list, but GovProp is ACTIVE_HEALTHY in production. Confirm tier."),
]
mkt = [
  ("Postiz", "$29.00 / month", 29.00, "Marketing", "Fixed", "Founder-provided"),
  ("Instantly.ai", "$99.00 / month", 99.00, "Marketing", "Fixed", "Founder-provided"),
  ("Apollo", "$59.00 / month", 59.00, "Marketing", "Fixed", "Founder-provided. May drop to $0 or rise to $159 — see scenario note below."),
  ("MillionVerifier", "$59.00 / month", 59.00, "Marketing", "Fixed", "Founder-provided"),
  ("Cold email domains (16)", "$141.00 first year", None, "Marketing", "Annual", "$141 year 1 (bought 2026-08-06); renews at $10.00/yr each = $160.00/yr."),
]
ANN_MAIN = None
ANN_CE = None
API_ROW = None
for name, billing, monthly, cat, status, note in ops:
    ws.cell(r, 1, name).font = BOLD
    ws.cell(r, 2, billing).font = BLACK
    if name == "Anthropic API":
        API_ROW = r          # monthly cell patched at the end, once API_MONTH_ROW is known
    elif name == "govhub.online (main domain)":
        ANN_MAIN = r
        c = ws.cell(r, 4, 40.00); c.font = BLUE; c.number_format = CUR2; c.border = BOX
        c = ws.cell(r, 3, f"=D{r}/12"); c.font = BLACK; c.number_format = CUR2
    else:
        c = ws.cell(r, 3, monthly); c.font = BLUE; c.number_format = CUR2; c.border = BOX
    if name not in ("govhub.online (main domain)",):
        c = ws.cell(r, 4, f"=C{r}*12"); c.font = BLACK; c.number_format = CUR2
    ws.cell(r, 5, cat).font = SMALL
    c = ws.cell(r, 6, status); c.font = WARN if status == "OPEN ITEM" else SMALL
    if status == "OPEN ITEM":
        c.fill = FILL_RED
    ws.cell(r, 7, note).font = SMALL
    r += 1
OPS_BOT = r - 1
c = ws.cell(r, 1, "Subtotal — Operational"); c.font = BOLD
c = ws.cell(r, 3, f"=SUM(C{STACK_TOP}:C{OPS_BOT})"); c.font = BOLD; c.number_format = CUR2; c.fill = FILL_GREY
c = ws.cell(r, 4, f"=SUM(D{STACK_TOP}:D{OPS_BOT})"); c.font = BOLD; c.number_format = CUR2; c.fill = FILL_GREY
OPS_SUB = r
r += 2

MKT_TOP = r
for name, billing, monthly, cat, status, note in mkt:
    ws.cell(r, 1, name).font = BOLD
    ws.cell(r, 2, billing).font = BLACK
    if name.startswith("Cold email domains"):
        ANN_CE = r
        c = ws.cell(r, 4, 141.00); c.font = BLUE; c.number_format = CUR2; c.border = BOX
        c = ws.cell(r, 3, f"=D{r}/12"); c.font = BLACK; c.number_format = CUR2
    else:
        c = ws.cell(r, 3, monthly); c.font = BLUE; c.number_format = CUR2; c.border = BOX
        c = ws.cell(r, 4, f"=C{r}*12"); c.font = BLACK; c.number_format = CUR2
    ws.cell(r, 5, cat).font = SMALL
    ws.cell(r, 6, status).font = SMALL
    ws.cell(r, 7, note).font = SMALL
    r += 1
MKT_BOT = r - 1
c = ws.cell(r, 1, "Subtotal — Marketing"); c.font = BOLD
c = ws.cell(r, 3, f"=SUM(C{MKT_TOP}:C{MKT_BOT})"); c.font = BOLD; c.number_format = CUR2; c.fill = FILL_GREY
c = ws.cell(r, 4, f"=SUM(D{MKT_TOP}:D{MKT_BOT})"); c.font = BOLD; c.number_format = CUR2; c.fill = FILL_GREY
MKT_SUB = r
r += 2

TOTAL_ALL = r
c = ws.cell(r, 1, "TOTAL MONTHLY BURN (run-rate)"); c.font = BOLD
c = ws.cell(r, 3, f"=C{OPS_SUB}+C{MKT_SUB}"); c.font = BOLD; c.number_format = CUR2
c.fill = FILL_YEL; c.border = BOX
c = ws.cell(r, 4, f"=D{OPS_SUB}+D{MKT_SUB}"); c.font = BOLD; c.number_format = CUR2
c.fill = FILL_YEL; c.border = BOX
ws.cell(r, 7, "Operational + Marketing. Anthropic API is already inside the Operational subtotal "
              "at the trailing month's actual — it is not added again here.").font = NOTE
TOTAL_FIXED = TOTAL_ALL
r += 1
c = ws.cell(r, 1, "Cash cost excluding founder time"); c.font = NOTE
ws.cell(r, 3, "").font = NOTE
ws.cell(r, 7, "No salary, contractor or founder-time cost is booked anywhere in this workbook. "
              "See the CAC warning on the Calculator tab before quoting any CAC or LTV:CAC number.").font = WARN
r += 2

r = band(ws, r, "ANNUAL RENEWAL CALENDAR", 7)
for i, h in enumerate(["Item", "Amount", "Next due", "Then", "Note"]):
    c = ws.cell(r, i+1, h); c.font = BOLD; c.fill = FILL_GREY
r += 1
for item, amt, due, then, note in [
  ("govhub.online", 40.00, "2026-09-26", "$40.00 every 9/26", "Main product domain"),
  ("16 cold email domains", 160.00, "2027-08-06", "$160.00/yr", "$141.00 was the year-1 promo; renewal is $10.00 × 16"),
]:
    ws.cell(r, 1, item).font = BOLD
    c = ws.cell(r, 2, amt); c.font = BLUE; c.number_format = CUR2
    ws.cell(r, 3, due).font = BLACK
    ws.cell(r, 4, then).font = BLACK
    ws.cell(r, 5, note).font = SMALL
    r += 1
r += 1

r = band(ws, r, "APOLLO SCENARIO — the one cost line you flagged as uncertain", 7)
for i, h in enumerate(["Scenario", "Apollo $/mo", "Total burn $/mo", "Δ vs today", "Note"]):
    c = ws.cell(r, i+1, h); c.font = BOLD; c.fill = FILL_GREY
r += 1
APOLLO_ROW = MKT_TOP + 2
for scen, val, note in [
  ("Today (as instructed)", 59.00, "Current run-rate"),
  ("Apollo dropped", 0.00, "You find a cheaper email-finding source"),
  ("Apollo +$100", 159.00, "No alternative found, tier upgrade required"),
]:
    ws.cell(r, 1, scen).font = BOLD
    c = ws.cell(r, 2, val); c.font = BLUE; c.number_format = CUR2; c.border = BOX
    c = ws.cell(r, 3, f"=$C${TOTAL_ALL}-$C${APOLLO_ROW}+B{r}"); c.font = BLACK; c.number_format = CUR2
    c = ws.cell(r, 4, f"=C{r}-$C${TOTAL_ALL}"); c.font = BLACK; c.number_format = CUR2
    ws.cell(r, 5, note).font = SMALL
    r += 1
r += 2

r = band(ws, r, "MONTHLY ACTUALS", 16)
ws.cell(r, 1, "Only Anthropic API has a queried actual per month. The subscription rows are seeded at the current "
              "run-rate for Aug-26 — fill earlier months from card statements if you want a true historical burn.").font = NOTE
r += 1
EXP_HDR = r
r = monthhdr(ws, r, "Cost line", "Source")
API_MONTH_ROW = r
r = datarow(ws, r, "Anthropic API", "v2_token_usage.est_cost_usd", API_COST, CUR2)
r = datarow(ws, r, "Claude subscription", "Founder-provided run-rate", [None]*11+[100.00], CUR2)
r = datarow(ws, r, "Resend", "Founder-provided run-rate", [None]*11+[20.00], CUR2)
r = datarow(ws, r, "Cloudflare", "Founder-provided run-rate", [None]*11+[5.00], CUR2)
r = datarow(ws, r, "Postiz", "Founder-provided run-rate", [None]*11+[29.00], CUR2)
r = datarow(ws, r, "Instantly.ai", "Founder-provided run-rate", [None]*11+[99.00], CUR2)
r = datarow(ws, r, "Apollo", "Founder-provided run-rate", [None]*11+[59.00], CUR2)
r = datarow(ws, r, "MillionVerifier", "Founder-provided run-rate", [None]*11+[59.00], CUR2)
r = datarow(ws, r, "Domains (16 cold email)", "Purchased 2026-08-06", [None]*11+[141.00], CUR2)
r = datarow(ws, r, "Domain (govhub.online)", "Renews 9/26", [None]*12, CUR2)
EXP_LAST = r - 1
c = ws.cell(r, 1, "TOTAL"); c.font = BOLD
ws.cell(r, 2, "").font = NOTE
for i in range(NM):
    cc = ws.cell(r, C0+i, f"=SUM({CL[i]}{API_MONTH_ROW}:{CL[i]}{EXP_LAST})")
    cc.font = BOLD; cc.number_format = CUR2; cc.fill = FILL_GREY
cc = ws.cell(r, C0+NM, f"=SUM({CL[0]}{r}:{CL[-1]}{r})")
cc.font = BOLD; cc.number_format = CUR2; cc.fill = FILL_GREY


# ============================================================== 9. UNIT ECONOMICS
ws = sheet("Unit Economics")
title(ws, "UNIT ECONOMICS",
      "Nearly every number here is blocked on the same thing: a first paying customer. This tab says so explicitly.", 8)
for col, w in zip("ABCDEFGH", [36, 18, 18, 22, 16, 16, 16, 50]):
    ws.column_dimensions[col].width = w
r = 4
for i, h in enumerate(["Metric", "Actual today", "At Calculator's assumptions", "Benchmark", "Status", "", "", "What it needs to become real"]):
    c = ws.cell(r, i+1, h); c.font = BOLD; c.fill = FILL_GREY
r += 1
ue = [
  ("New customers / mo", 0, f"=Calculator!B{TOT_CUST}", "—", NUM1,
   "Actual is 0 because there is no billing path."),
  ("Effective ARPA", 0, f"=Calculator!B{EFF_ARPA}", "—", CUR,
   "Modelled from the live price list × an assumed tier mix."),
  ("Total S&M spend / mo", "=Expenses!C30", f"=Calculator!B{TOT_CUST+2}", "—", CUR,
   "Real: the marketing subtotal from the Expenses tab."),
  ("Blended CAC", "n/a", f"=Calculator!B{BCAC}", "Judge vs LTV only", CUR,
   "Needs paying customers in the denominator. Spend is known; customers are 0."),
  ("Gross margin %", "n/a", f"=Calculator!B{GM}", "75–85%", PCT,
   "API cost is trivially small, so 85% is a defensible modelling assumption."),
  ("Monthly churn %", "n/a", f"=Calculator!B{CHURN}", "SMB 3–7%", PCT,
   "Needs one paying cohort with 60+ days of history."),
  ("LTV (GM-adjusted)", "n/a", f"=Calculator!B{LTV}", "—", CUR,
   "ARPA × GM% ÷ churn. Blocked entirely on a real churn rate."),
  ("LTV : CAC", "n/a", f"=Calculator!B{TOT_CUST+7}", "≥3x healthy, 5x+ strong", MULT,
   "Blocked on both LTV and CAC."),
  ("CAC payback (months)", "n/a", f"=Calculator!B{TOT_CUST+8}", "≤12 months", NUM1,
   "Annual prepay would shorten this materially — 20% annual share is already modelled."),
  ("Steady-state MRR ceiling", "n/a", f"=Calculator!B{CEIL}", "—", CUR,
   "New MRR ÷ churn. The number this engine converges to. Worth knowing before you scale spend."),
]
for label, actual, modelled, bench, fmt, need in ue:
    ws.cell(r, 1, label).font = BOLD
    c = ws.cell(r, 2, actual)
    c.font = BLUE if isinstance(actual, (int, float)) else (GREEN if str(actual).startswith("=") else NOTE)
    if isinstance(actual, (int, float)) or str(actual).startswith("="):
        c.number_format = fmt
    c.alignment = Alignment(horizontal="center")
    c = ws.cell(r, 3, modelled); c.font = GREEN; c.number_format = fmt; c.border = BOX
    ws.cell(r, 4, bench).font = SMALL
    c = ws.cell(r, 5, "MODELLED" if actual == "n/a" else "REAL")
    c.font = BOLD; c.alignment = Alignment(horizontal="center")
    c.fill = FILL_YEL if actual == "n/a" else FILL_GRN
    ws.cell(r, 8, need).font = SMALL
    r += 1
r += 1
ws.cell(r, 1, "Column B is what you can defend with data today. Column C is what the model produces at the "
              "Calculator's assumptions. Never present column C as though it were column B — that is exactly the "
              "mistake the previous version of this workbook made.").font = NOTE


# ============================================================== 10. PROJECTIONS
ws = sheet("Projections")
title(ws, "PROJECTIONS — 24-month driver model (Sep-26 → Aug-28)",
      "Every driver is a live link to the Calculator tab. Change levers there; this tab follows.", 27)
r = 4
PMONTHS = []
_y, _m = 2026, 9
for _ in range(24):
    PMONTHS.append(f"{['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'][_m-1]}-{str(_y)[2:]}")
    _m += 1
    if _m > 12:
        _m = 1; _y += 1
PC0 = 3
PCL = [get_column_letter(PC0+i) for i in range(24)]

ws.cell(r, 1, "Cold email is modelled as launching in Sep-26 (first full month after the 2026-08-27 earliest "
              "defensible start). SEO clicks compound at the growth rate below.").font = NOTE
r += 2
r = band(ws, r, "DRIVERS (live from Calculator)", 27)
PD = r
drivers = [
  ("Organic sessions, month 1", f"=Calculator!B{S_SESS}", NUM),
  ("Monthly session growth %", 0.15, PCT),
  ("Session → signup %", f"=Calculator!B{S_S2S}", PCT2),
  ("Signup → activated %", f"=Calculator!B{S_S2A}", PCT),
  ("Activated → paid %", f"=Calculator!B{S_A2P}", PCT),
  ("Emails sent / mo (at full ramp)", f"=Calculator!B{E_SENT}", NUM),
  ("Email sends growth %", 0.05, PCT),
  ("Email → customer conversion %", f"=IFERROR(Calculator!B{E_CUST}/Calculator!B{E_SENT},0)", PCT2),
  ("Direct customers / mo", f"=Calculator!B{O_CUST}", NUM1),
  ("Effective ARPA", f"=Calculator!B{EFF_ARPA}", CUR),
  ("Monthly churn %", f"=Calculator!B{CHURN}", PCT),
  ("Monthly expansion %", f"=Calculator!B{EXPAN}", PCT),
  ("Gross margin %", f"=Calculator!B{GM}", PCT),
  ("Fixed cost / mo", f"=Expenses!C{TOTAL_ALL}", CUR2),
  ("Monthly cost growth %", 0.02, PCT),
]
for label, val, fmt in drivers:
    ws.cell(r, 1, label).font = BLACK
    c = ws.cell(r, 2, val)
    c.font = GREEN if str(val).startswith("=") else BLUE
    c.number_format = fmt; c.border = BOX
    if not str(val).startswith("="):
        c.fill = FILL_YEL
    r += 1
D_SESS, D_SGROW, D_S2S, D_S2A, D_A2P = PD, PD+1, PD+2, PD+3, PD+4
D_EM, D_EGROW, D_E2C, D_DIR, D_ARPA = PD+5, PD+6, PD+7, PD+8, PD+9
D_CHURN, D_EXP, D_GM, D_FIX, D_CGROW = PD+10, PD+11, PD+12, PD+13, PD+14
r += 1

r = band(ws, r, "MODEL", 27)
PH = r
ws.cell(r, 1, "Line").font = BOLD
ws.cell(r, 2, "").font = BOLD
for i, m in enumerate(PMONTHS):
    c = ws.cell(r, PC0+i, m); c.font = BOLD; c.fill = FILL_GREY
    c.alignment = Alignment(horizontal="center")
ws.column_dimensions['A'].width = 34
ws.column_dimensions['B'].width = 4
for i in range(24):
    ws.column_dimensions[PCL[i]].width = 10.5
r += 1


def prow(row, label, first, rest, fmt=NUM1, bold=False):
    c = ws.cell(row, 1, label); c.font = BOLD if bold else BLACK
    c.alignment = Alignment(indent=0 if bold else 1)
    ws.cell(row, PC0, first).number_format = fmt
    ws.cell(row, PC0).font = BLACK
    for i in range(1, 24):
        cc = ws.cell(row, PC0+i, rest.format(p=PCL[i-1], c=PCL[i]))
        cc.font = BLACK; cc.number_format = fmt
    return row + 1


P_SESS = r
r = prow(r, "Organic sessions", f"=$B${D_SESS}", "={p}%d*(1+$B$%d)" % (P_SESS, D_SGROW), NUM)
P_SSIGN = r
r = prow(r, "Organic signups", f"=C{P_SESS}*$B${D_S2S}", "={c}%d*$B$%d" % (P_SESS, D_S2S))
P_SACT = r
r = prow(r, "Organic activated", f"=C{P_SSIGN}*$B${D_S2A}", "={c}%d*$B$%d" % (P_SSIGN, D_S2A))
P_SCUST = r
r = prow(r, "New customers — SEO", f"=C{P_SACT}*$B${D_A2P}", "={c}%d*$B$%d" % (P_SACT, D_A2P))
P_EM = r
r = prow(r, "Emails sent", f"=$B${D_EM}", "={p}%d*(1+$B$%d)" % (P_EM, D_EGROW), NUM)
P_ECUST = r
r = prow(r, "New customers — cold email", f"=C{P_EM}*$B${D_E2C}", "={c}%d*$B$%d" % (P_EM, D_E2C))
P_OCUST = r
r = prow(r, "New customers — direct", f"=$B${D_DIR}", "=$B$%d" % D_DIR)
P_TCUST = r
r = prow(r, "Total new customers", f"=C{P_SCUST}+C{P_ECUST}+C{P_OCUST}",
         "={c}%d+{c}%d+{c}%d" % (P_SCUST, P_ECUST, P_OCUST), NUM1, bold=True)
r += 1
P_BEGC = r
r = prow(r, "Beginning customers", "=0", "={p}%d" % (r+2), NUM1)
P_CHRN = r
r = prow(r, "Churned customers", f"=C{P_BEGC}*$B${D_CHURN}", "={c}%d*$B$%d" % (P_BEGC, D_CHURN))
P_ENDC = r
r = prow(r, "Ending customers", f"=C{P_BEGC}+C{P_TCUST}-C{P_CHRN}",
         "={c}%d+{c}%d-{c}%d" % (P_BEGC, P_TCUST, P_CHRN), NUM1, bold=True)
r += 1
P_BEGM = r
r = prow(r, "Beginning MRR", "=0", "={p}%d" % (r+5), CUR)
P_NEWM = r
r = prow(r, "New MRR", f"=C{P_TCUST}*$B${D_ARPA}", "={c}%d*$B$%d" % (P_TCUST, D_ARPA), CUR)
P_EXPM = r
r = prow(r, "Expansion MRR", f"=C{P_BEGM}*$B${D_EXP}", "={c}%d*$B$%d" % (P_BEGM, D_EXP), CUR)
P_CHRM = r
r = prow(r, "Churned MRR", f"=C{P_BEGM}*$B${D_CHURN}", "={c}%d*$B$%d" % (P_BEGM, D_CHURN), CUR)
P_NETM = r
r = prow(r, "Net new MRR", f"=C{P_NEWM}+C{P_EXPM}-C{P_CHRM}",
         "={c}%d+{c}%d-{c}%d" % (P_NEWM, P_EXPM, P_CHRM), CUR)
P_ENDM = r
r = prow(r, "Ending MRR", f"=C{P_BEGM}+C{P_NETM}", "={c}%d+{c}%d" % (P_BEGM, P_NETM), CUR, bold=True)
P_ARR = r
r = prow(r, "ARR (run-rate)", f"=C{P_ENDM}*12", "={c}%d*12" % P_ENDM, CUR)
r += 1
P_GP = r
r = prow(r, "Gross profit", f"=C{P_ENDM}*$B${D_GM}", "={c}%d*$B$%d" % (P_ENDM, D_GM), CUR)
P_COST = r
r = prow(r, "Total cost", f"=$B${D_FIX}", "={p}%d*(1+$B$%d)" % (P_COST, D_CGROW), CUR)
P_NI = r
r = prow(r, "Net income / (burn)", f"=C{P_GP}-C{P_COST}", "={c}%d-{c}%d" % (P_GP, P_COST), CUR, bold=True)
r += 2

r = band(ws, r, "SCENARIO SUMMARY", 27)
for label, f, fmt in [
  ("MRR — month 12 (Aug-27)", f"=N{P_ENDM}", CUR),
  ("MRR — month 24 (Aug-28)", f"=Z{P_ENDM}", CUR),
  ("ARR — month 24", f"=Z{P_ARR}", CUR),
  ("Customers — month 24", f"=Z{P_ENDC}", NUM1),
  ("Ceiling if acquisition never grows", f"=Calculator!B{CEIL}", CUR),
  ("First cash-flow-positive month", f'=IFERROR(INDEX($C${PH}:$Z${PH},MATCH(TRUE,INDEX($C${P_NI}:$Z${P_NI}>0,0),0)),"not within 24 months")', None),
  ("Months profitable (of 24)", f"=COUNTIF(C{P_NI}:Z{P_NI},\">0\")", NUM),
]:
    ws.cell(r, 1, label).font = BOLD
    c = ws.cell(r, 2, f); c.font = GREEN if "Calculator" in str(f) else BLACK
    if fmt:
        c.number_format = fmt
    c.fill = FILL_GRN; c.border = BOX
    r += 1
r += 1
ws.cell(r, 1, "Why month-24 MRR exceeds the 'ceiling' above: the ceiling is the level this engine converges to if "
              "acquisition NEVER grows. This model grows organic sessions 15%/mo and email volume 5%/mo, so it climbs "
              "past it. The ceiling is the useful number for 'what happens if we stop improving' — the model is the "
              "number for 'what happens if we keep pushing'. Show an investor both.").font = NOTE
r += 2
ws.cell(r, 1, "⚠ These projections assume Stripe billing ships and cold email launches. Neither has happened yet. "
              "They are a plan, not a forecast, and should be presented that way.").font = WARN
ws.cell(r, 1).fill = FILL_RED


# ============================================================== 11. BENCHMARKS
ws = sheet("Benchmarks")
title(ws, "BENCHMARKS & TARGETS — what good looks like",
      "Compiled from David Skok (forEntrepreneurs/Matrix), David Sacks (Craft Ventures), OpenView, and 2026 cold-email datasets.", 5)
for col, w in zip("ABCDE", [18, 40, 30, 66, 10]):
    ws.column_dimensions[col].width = w
r = 4
for i, h in enumerate(["Category", "Metric", "Benchmark / target", "Why it matters / how it applies to GovHub"]):
    c = ws.cell(r, i+1, h); c.font = BOLD; c.fill = FILL_GREY
r += 1
bm = [
 ("Growth","MRR growth (MoM)","10%+ strong pre-$1M ARR","Best single proof of product-market fit. Not measurable at $0."),
 ("Growth","Net New MRR mix","New + expansion > churn + contraction","Track the four-part waterfall from the first paying month."),
 ("Retention","Logo churn (monthly)","SMB SaaS 3–7%/mo; aim <5%","Govcon has bid-cycle seasonality — expect lumpy early churn."),
 ("Retention","Net revenue retention","≥100%","Solo→Pro upgrades and Team seats are the expansion path."),
 ("Retention","Quick ratio","≥4x","(New+Expansion) ÷ (Churned+Contraction). <2 = leaky bucket."),
 ("Unit economics","LTV : CAC","≥3x healthy; 5x+ strong","Skok/Craft standard. <3x means do not scale spend yet."),
 ("Unit economics","CAC payback","≤12 months","Annual prepay (10× monthly) shortens effective payback materially."),
 ("Unit economics","Gross margin","75–85%","Your API cost is ~$27/mo at peak — margin is not your problem."),
 ("Unit economics","Magic number",">0.75 = step on gas; <0.5 = fix first","Net-new ARR ÷ prior-period S&M."),
 ("Unit economics","Rule of 40","Growth% + margin% ≥ 40","Matters past $1M ARR; directional before that."),
 ("Funnel","Trial → paid (no card)","~10–25% typical band","The Calculator assumes 12% organic / 30% email. Both are assumptions."),
 ("Funnel","Visitor → trial","0.5–2% of unique visitors","Landing-page and offer clarity is the lever."),
 ("Funnel","Activation rate","40–60% of signups","GovHub's actual is 70% (21/30) — genuinely above benchmark."),
 ("Cold email","Reply rate (of delivered)","2026 avg ~3.4%; 5%+ good","Declining industry-wide; segmentation is the biggest lever."),
 ("Cold email","Positive reply rate","~1–2% solid","Track positive separately — total replies include 'remove me'."),
 ("Cold email","Meeting booked rate","~0.5–2.5% of delivered","Varies with offer and follow-up quality."),
 ("Cold email","Bounce rate","<2%","MillionVerifier at $59/mo is the control for this."),
 ("Cold email","Spam complaint rate","<0.3% (Gmail/Yahoo mandate)","Hard requirement for bulk senders since 2024."),
 ("Cold email","Volume per mailbox","50–100/day AFTER months of history","Your domains are days old. 15/day is the right starting cap."),
 ("Cold email","Warm-up before sending","2 weeks minimum; 3 weeks holds up","Domains created 2026-08-06 → 2026-08-27 is day 21."),
 ("Cold email","Follow-ups","2–3 touches; ~40%+ of replies come from follow-ups","Most senders quit after one email."),
 ("SEO","Time to compounding","6–12 months","You are at month 2 of real indexation. Judge on keywords, not revenue."),
 ("SEO","CTR by position","Position 1–3 ≈ 10–30%+; page 2+ ≈ ~0","'govwin' at position 23 earns almost nothing. Moving it is the whole game."),
 ("Cash","Runway","≥12 months","At ~$413/mo burn this is governed by your personal runway, not the company's."),
]
for cat, metric, bench, why in bm:
    ws.cell(r, 1, cat).font = BOLD
    ws.cell(r, 2, metric).font = BLACK
    ws.cell(r, 3, bench).font = BLACK
    ws.cell(r, 4, why).font = SMALL
    r += 1

# -------------------------------------------------- patch forward references
LASTC, PREVC = CL[-1], CL[-2]          # N, M

# Expenses cost-stack: Anthropic API monthly = trailing-month actual
xe = wb["Expenses"]
c = xe.cell(API_ROW, 3, f"={LASTC}{API_MONTH_ROW}")
c.font = GREEN; c.number_format = CUR2
c = xe.cell(API_ROW, 4, f"=C{API_ROW}*12"); c.font = BLACK; c.number_format = CUR2

# Dashboard: replace every guessed cross-sheet reference with the real address
xd = wb["Dashboard"]
DASH_REFS = {
    "MRR":                                 (f"=Revenue!{TOTC}{R_END}",        f"=Revenue!{PREVC}{R_END}"),
    "Paying customers":                    (f"=Revenue!{TOTC}{R_NEW+2}",      f"=Revenue!{PREVC}{R_NEW+2}"),
    "Total signups (all-time)":            (f"=Funnel!C{F_TOP+2}",            None),
    "External signups (all-time)":         (f"=Funnel!C{F_TOP+3}",            None),
    "Active trials right now":             (f"=Funnel!B{T_TOP+2}",            None),
    "Organic keywords ranking (top 100)":  (f"=SEO!{TOTC}{SEO_KW_ROW}",       f"=SEO!{CL[9]}{SEO_KW_ROW}"),
    "Est. monthly organic sessions":       (f"=SEO!{TOTC}{SEO_KW_ROW+1}",     f"=SEO!{CL[9]}{SEO_KW_ROW+1}"),
    "Proposals generated (v2_submissions)":(f"=Funnel!C{F_TOP+5}",            None),
    "Anthropic API spend / mo":            (f"=Expenses!{LASTC}{API_MONTH_ROW}", f"=Expenses!{PREVC}{API_MONTH_ROW}"),
    "Cold emails sent":                    (f"='Cold Email'!{TOTC}{CE_SENT_ROW}", None),
}
for rr in range(1, xd.max_row + 1):
    lbl = xd.cell(rr, 1).value
    if lbl in DASH_REFS:
        cur, prior = DASH_REFS[lbl]
        xd.cell(rr, 2).value = cur
        if prior:
            xd.cell(rr, 3).value = prior
# ARR row sits directly under Paying customers' MRR row — rebuild from the MRR row
for rr in range(1, xd.max_row + 1):
    if xd.cell(rr, 1).value == "MRR":
        mrr_row = rr
    if xd.cell(rr, 1).value == "ARR (run-rate)":
        xd.cell(rr, 2).value = f"=B{mrr_row}*12"
        xd.cell(rr, 3).value = f"=C{mrr_row}*12"
        break

# Calculator + Unit Economics: point cold-email spend / S&M at the Marketing subtotal
xc = wb["Calculator"]
xc.cell(SP_EMAIL, 2).value = f"=Expenses!C{MKT_SUB}"
xc.cell(SP_EMAIL, 2).font = GREEN
xc.cell(SP_EMAIL, 2).number_format = CUR
xc.cell(MBRK, 2).value = (
    f'=IF(B{CEIL}*B{GM}<=(Expenses!$C${TOTAL_ALL}+B{SP_LABOUR}),"never",'
    f'LN(1-((Expenses!$C${TOTAL_ALL}+B{SP_LABOUR})/B{GM})*B{CHURN}/B{NEWMRR})/LN(1-B{CHURN}))')
# Projections fixed-cost driver: cash burn + imputed founder time
wb["Projections"].cell(D_FIX, 2).value = f"=Expenses!C{TOTAL_ALL}+Calculator!B{SP_LABOUR}"
xu = wb["Unit Economics"]
for rr in range(1, xu.max_row + 1):
    if xu.cell(rr, 1).value == "Total S&M spend / mo":
        xu.cell(rr, 2).value = f"=Expenses!C{MKT_SUB}"
        break

# (Projections fixed-cost driver is set above: cash burn + imputed founder time)

del wb["Sheet"]
wb.save(OUT)
print("saved", OUT)
print("Calculator refs: ARPA row", ARPA, "EFF_ARPA", EFF_ARPA, "TOT_CUST", TOT_CUST,
      "CEIL", CEIL, "M10K", M10K, "MBRK", MBRK, "BCAC", BCAC, "LTV", LTV)
print("Expenses: OPS_SUB", OPS_SUB, "MKT_SUB", MKT_SUB, "TOTAL_FIXED", TOTAL_FIXED,
      "TOTAL_ALL", TOTAL_ALL, "API_MONTH_ROW", API_MONTH_ROW, "APOLLO_ROW", APOLLO_ROW)
